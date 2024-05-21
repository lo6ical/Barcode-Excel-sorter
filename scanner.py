import click
from openpyxl import load_workbook, Workbook
import threading

# Function to handle barcode input
def handle_input(wb, excel_file):
    # Infinite loop to handle barcode inputs
    while True:
        barcode = input("Scan a barcode")  # Prompt user to scan a barcode
        length = len(barcode)  # Get the length of the scanned barcode
        column = 0

        # Check if barcode is empty
        if length == 0:
            click.echo("invalid barcode")  # Notify user of invalid barcode
            continue
        
        # Set column based on barcode input length
        if length == 9:
            column = 1
        if length == 14:
            column = 2
        
        try:
            ws = wb.active  # Get the active worksheet
            # Find the column to insert the barcode based on length
            col = ws.iter_cols(min_col=column, max_col=column)
            # Find the next empty row in the specified column
            row = next((cell.row for col in col for cell in col if cell.value is None), None)
            if row is None:
                row = ws.max_row + 1  # If no empty cell is found, add to the next row
            # Insert the barcode into the cell
            ws.cell(row=row, column=column, value=barcode)
            # Save the workbook
            wb.save(excel_file)
            click.echo(f"Barcode {barcode} inserted into column {column+1} of the Excel file")
        except Exception as e:
            click.echo(f"Error: {str(e)}")  # Handle any exceptions and notify the user

@click.command()
@click.option('--excel-file', default='<Excel file name>', help='<Path to Excel file')

def main(excel_file):
    click.echo("Barcode scanner program started. Press Ctrl+C to exit.")
    
    try:
        wb = load_workbook(excel_file)  # Try to load the existing Excel file
    except FileNotFoundError:
        click.echo("Excel file not found.")  # Notify user if file is not found
        wb = Workbook()  # Create a new Excel workbook if file doesn't exist
        wb.create_sheet(title='Sheet1', index=0)  # Create a new sheet in the workbook
        wb.save(excel_file)  # Save the newly created workbook
        wb = load_workbook(excel_file)  # Reload the newly created workbook

    # Create a thread for handling barcode input
    input_thread = threading.Thread(target=handle_input, args=(wb, excel_file), daemon=True)
    input_thread.start()

    # Main thread can continue with any other tasks or logging
    # Wait for the input thread to finish
    input_thread.join()

main()
